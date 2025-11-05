import pandas as pd
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from collections import deque
from typing import List, Dict, Set, Tuple, Any
import numpy as np
import pandas.api.types as pd_types
import json


# ======================================================================
# GIAI ĐOẠN 1: PHÁT HIỆN BẢNG (Giữ nguyên từ trước)
# ======================================================================

# --- BƯỚC 1: TẠO BẢN ĐỒ TRA CỨU Ô GỘP ---
def _create_merged_cell_map(ws: Worksheet) -> Dict[Tuple[int, int], Tuple[int, int]]:
    """
    Tạo một dict (map) để tra cứu ô "cha" (ô top-left chứa style)
    từ bất kỳ tọa độ ô "con" nào.
    
    Returns:
        Dict[(con_r, con_c), (cha_r, cha_c)]
    """
    merged_map = {}
    # Lặp qua tất cả các dải ô gộp trong sheet
    for merged_range in ws.merged_cells.ranges:
        # Lấy tọa độ (1-based index) của dải ô
        min_col, min_row, max_col, max_row = merged_range.bounds
        
        # Tọa độ ô "cha" (ô top-left)
        parent_coord = (min_row, min_col)
        
        # Lặp qua tất cả ô "con" trong dải (bao gồm cả ô cha)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                # Ánh xạ ô con về ô cha
                merged_map[(r, c)] = parent_coord
    return merged_map

# --- BƯỚC 2: TẠO BẢN ĐỒ NHIỆT BORDER (ĐÃ XỬ LÝ Ô GỘP) ---
def _create_border_heatmap(ws: Worksheet, merged_map: Dict) -> List[List[bool]]:
    """
    Tạo một bản đồ 2D (heatmap) của sheet.
    True = "Đất" (Ô này có border, hoặc là 1 phần của ô gộp có border)
    False = "Biển" (Ô này không có border)
    
    Bản đồ này sử dụng 0-based index để dễ dàng cho Bước 3.
    """
    max_r, max_c = ws.max_row, ws.max_column
    
    # Tạo bản đồ rỗng (0-indexed)
    # heatmap[hàng][cột]
    heatmap = [[False for _ in range(max_c)] for _ in range(max_r)]

    # Lặp qua từng ô trong sheet (1-indexed)
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            
            # 1. Tìm tọa độ ô chứa style
            # Mặc định là chính nó
            style_coord = (r, c)
            if (r, c) in merged_map:
                # Nếu là ô con, lấy tọa độ ô cha
                style_coord = merged_map[(r, c)]
            
            # 2. Lấy ô style từ tọa độ
            style_cell = ws.cell(row=style_coord[0], column=style_coord[1])
            
            # 3. Kiểm tra border của ô style
            b = style_cell.border
            # Chỉ cần 1 cạnh có style là coi như ô đó có border
            if (b.left.style or b.right.style or b.top.style or b.bottom.style):
                # Đánh dấu "Đất" (True) vào heatmap (0-indexed)
                heatmap[r-1][c-1] = True
                
    return heatmap

# --- BƯỚC 3: TÌM "CỤM BORDER" (BFS) ---
def _find_clusters(heatmap: List[List[bool]]) -> List[List[Tuple[int, int]]]:
    """
    Chạy thuật toán BFS (Breadth-First Search) trên heatmap
    
    Returns:
        List các cụm, mỗi cụm là 1 List các tọa độ (r, c) (0-indexed).
    """
    if not heatmap: 
        return []
        
    rows = len(heatmap)
    cols = len(heatmap[0])
    
    visited = set()  # Set chứa các tọa độ (r, c) (0-indexed) đã ghé thăm
    clusters = []    # List chứa các cụm

    for r in range(rows):
        for c in range(cols):
            # Nếu ô này là "Đất" (True) và chưa được ghé thăm
            if heatmap[r][c] and (r, c) not in visited:
                
                # Bắt đầu một cụm mới
                new_cluster = []
                q = deque([(r, c)]) # Hàng đợi cho BFS
                visited.add((r, c))

                while q:
                    curr_r, curr_c = q.popleft()
                    # Thêm tọa độ (0-indexed) vào cụm
                    new_cluster.append((curr_r, curr_c))

                    # Kiểm tra 4 hướng lân cận (trên, dưới, trái, phải)
                    for dr, dc in [(0, 1), (0, -1), (1, 0), (-1, 0)]:
                        nr, nc = curr_r + dr, curr_c + dc

                        # Kiểm tra xem có nằm trong ranh giới bản đồ không
                        if 0 <= nr < rows and 0 <= nc < cols:
                            # Nếu ô lân cận là "Đất" và chưa ghé thăm
                            if heatmap[nr][nc] and (nr, nc) not in visited:
                                visited.add((nr, nc))
                                q.append((nr, nc))
                
                # Sau khi while kết thúc, thêm cụm mới vào danh sách
                clusters.append(new_cluster)
                
    return clusters

# --- BƯỚC 4: LỌC CỤM & LẤY TỌA ĐỘ (BOUNDING BOX) ---
def _filter_and_get_boundaries(clusters: List[List[Tuple[int, int]]], 
                               min_width: int = 5, 
                               min_height: int = 3) -> List[Dict[str, int]]:
    """
    Lặp qua các cụm, lọc bỏ "nhiễu" (cụm quá nhỏ),
    và trả về tọa độ (bounding box) của các "bảng" hợp lệ.
    
    Tọa độ trả về là 1-indexed (để khớp với Excel).
    """
    final_table_boundaries = []
    
    for cluster in clusters:
        if not cluster:
            continue
            
        # Lấy tất cả tọa độ (0-indexed) của cụm
        all_r = [r for r, c in cluster]
        all_c = [c for r, c in cluster]
        
        # Tìm min/max (0-indexed)
        min_r, max_r = min(all_r), max(all_r)
        min_c, max_c = min(all_c), max(all_c)
        
        # Tính toán kích thước
        width = max_c - min_c + 1
        height = max_r - min_r + 1
        
        # Áp dụng bộ lọc (heuristic)
        if width >= min_width and height >= min_height:
            
            # Nếu đủ lớn, lưu lại tọa độ (chuyển về 1-indexed)
            final_table_boundaries.append({
                'min_row': min_r + 1,
                'max_row': max_r + 1,
                'min_col': min_c + 1,
                'max_col': max_c + 1
            })
            
    return final_table_boundaries

# --- HÀM TỔNG HỢP (MAIN FUNCTION) ---
def detect_tables(file_path: str, sheet_name: str, 
                  min_width: int = 5, 
                  min_height: int = 3) -> List[Dict[str, int]]:
    """
    Phát hiện tất cả các "bảng" (được định nghĩa bằng border)
    trong một sheet Excel.
    
    Args:
        file_path: Đường dẫn đến file Excel.
        sheet_name: Tên sheet cần xử lý.
        min_width: Chiều rộng tối thiểu để coi là 1 bảng (ý tưởng "line > 5").
        min_height: Chiều cao tối thiểu để coi là 1 bảng.
        
    Returns:
        Một list các dict, mỗi dict chứa tọa độ 1-indexed của bảng.
        Ví dụ: [{'min_row': 2, 'max_row': 12, 'min_col': 1, 'max_col': 26}]
    """
    try:
        # data_only=True để đọc giá trị (nếu cần), không phải công thức
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            print(f"Lỗi: Không tìm thấy sheet '{sheet_name}' trong file.")
            return []
        ws = wb[sheet_name]
    except Exception as e:
        print(f"Lỗi khi tải file hoặc sheet: {e}")
        return []

    # --- Chạy 4 bước của Giai đoạn 1 ---
    
    # Bước 1:
    print(f"Bước 1: Đang tạo bản đồ ô gộp...")
    merged_map = _create_merged_cell_map(ws)
    print(f"Bước 1: Hoàn thành. Tìm thấy {len(merged_map)} ô con trong các ô gộp.")
    
    # Bước 2:
    print(f"Bước 2: Đang tạo bản đồ nhiệt border (có xử lý ô gộp)...")
    heatmap = _create_border_heatmap(ws, merged_map)
    print("heatmap",heatmap)
    print("Bước 2: Hoàn thành.")
    
    # Bước 3:
    print(f"Bước 3: Đang tìm các cụm border...")
    clusters = _find_clusters(heatmap)
    print("clusters",clusters)
    print(f"Bước 3: Hoàn thành. Tìm thấy {len(clusters)} cụm.")
    
    # Bước 4:
    print(f"Bước 4: Đang lọc cụm và lấy tọa độ (min_width={min_width}, min_height={min_height})...")
    boundaries = _filter_and_get_boundaries(clusters, min_width, min_height)
    print(f"Bước 4: Hoàn thành. Tìm thấy {len(boundaries)} bảng hợp lệ.")
    
    wb.close()
    return boundaries







def debug_extract_data(file_path: str, sheet_name: str, 
                       boundary: Dict[str, int]) -> pd.DataFrame:
    """
    Đọc và trả về dữ liệu thô (raw data) từ BÊN TRONG một tọa độ (boundary)
    đã được phát hiện, dùng cho mục đích kiểm tra (debug).
    
    Tọa độ boundary nhận vào là 1-indexed.
    """
    
    # 1. Chuyển đổi tọa độ 1-indexed (từ detect_tables) 
    #    sang 0-indexed (cho pandas)
    
    # Hàng 3 (1-indexed) -> skiprows=2 (bỏ qua hàng 0, 1)
    skip_rows = boundary['min_row'] - 1
    
    # Số hàng cần đọc
    num_rows = boundary['max_row'] - boundary['min_row'] + 1
    
    # Cột 1 (1-indexed) -> cột 0 (0-indexed)
    # Cột 26 (1-indexed) -> cột 25 (0-indexed)
    # Chúng ta cần list [0, 1, ..., 25]
    cols_to_use = list(range(
        boundary['min_col'] - 1,  # (1-1) = 0
        boundary['max_col']       # (26) -> range() sẽ dừng ở 25
    ))
    
    if not cols_to_use:
        print("Lỗi: Không có cột nào để đọc.")
        return pd.DataFrame()

    # 2. Đọc file Excel chỉ trong phạm vi đã định
    try:
        raw_table_df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=None,        # Không giả định header, đọc thô
            skiprows=skip_rows,   # Bỏ qua các hàng bên trên
            nrows=num_rows,     # Chỉ đọc số hàng của bảng
            usecols=cols_to_use   # Chỉ đọc các cột của bảng
        )
        
        # Đặt lại index cột để dễ nhìn (0, 1, 2...)
        raw_table_df.columns = range(raw_table_df.shape[1])
        
        return raw_table_df
        
    except Exception as e:
        print(f"Lỗi khi trích xuất dữ liệu debug: {e}")
        return pd.DataFrame()
    

# ======================================================================
# GIAI ĐOẠN 2: TRÍCH XUẤT JSON (Code mới)
# ======================================================================
import pandas as pd
import json
from typing import Dict, List, Any, Optional, Tuple
import re

# --- [GIAI ĐOẠN 2: PARSE LOGIC - HÀM MỚI] ---


def detect_header_split_point(
    raw_table_df: pd.DataFrame, 
    worksheet: Worksheet,
    boundary: Dict[str, int],
    merged_map: Dict[Tuple[int, int], Tuple[int, int]],
    border_threshold: float = 0.95 # <-- ĐÃ THAY ĐỔI GIÁ TRỊ MẶC ĐỊNH
) -> int:
    """
    (Phiên bản V3 - ĐÃ SỬA BUG Ô GỘP)
    Phát hiện header bằng cách quét "KHÔNG GIAN" (ranh giới) giữa các hàng.
    
    Logic:
    1. Quét từng "ranh giới" (N-1 ranh giới cho N hàng).
    2. Nếu ranh giới NẰM BÊN TRONG một ô gộp, bỏ qua (coi như không có border).
    3. Nếu là ranh giới THỰC SỰ, kiểm tra CẢ hai "bờ":
       - `border.bottom` (kẻ dưới) của HÀNG TRÊN.
       - `border.top` (kẻ trên) của HÀNG DƯỚI.
    4. Tìm ranh giới "ỨNG VIÊN" CUỐI CÙNG thỏa mãn threshold.
    """
    
    # --- (Phần Validate và lấy total_cols, total_rows giữ nguyên) ---
    if not 0 <= border_threshold <= 1:
        print(f"⚠ CẢNH BÁO: border_threshold phải từ 0.0 đến 1.0, nhận được: {border_threshold}")
        border_threshold = max(0.0, min(1.0, border_threshold))
    
    total_columns = raw_table_df.shape[1]
    if total_columns == 0:
        return -1

    total_rows = raw_table_df.shape[0]
    if total_rows <= 1:
        return -1

    print(f"\n[detect_header_split_point] Quét {total_rows - 1} ranh giới, {total_columns} cột")
    print(f"  Threshold: {border_threshold} ({border_threshold*100:.1f}%)")
    print(f"  Số cells tối thiểu: {int(border_threshold * total_columns)}/{total_columns}\n")

    last_split_row_idx = -1 
    
    # Quét N-1 ranh giới
    for r_idx in range(total_rows - 1):
        
        real_row_above = boundary['min_row'] + r_idx
        real_row_below = boundary['min_row'] + r_idx + 1
        print(r_idx, real_row_above, real_row_below)
        horizontal_count = 0
        
        # Quét từ trái qua phải
        for c_idx in range(boundary['min_col'], boundary['max_col'] + 1):
            
            # 1. Tìm "ô cha" (ô chứa style) cho cả hai
            coord_above = (real_row_above, c_idx)
            style_coord_above = merged_map.get(coord_above, coord_above)
            
            coord_below = (real_row_below, c_idx)
            style_coord_below = merged_map.get(coord_below, coord_below)
            
            has_bottom_border = False
            has_top_border = False

            # --- (PHẦN SỬA LỖI LOGIC QUAN TRỌNG) ---
            if style_coord_above == style_coord_below:
                # Nếu "ô cha" của cả hai LÀ MỘT
                # (Ví dụ: A3 và A4 cùng có cha là A3)
                # -> Đây là ranh giới "ảo" BÊN TRONG một ô gộp.
                # -> Bỏ qua, coi như không có border.
                pass
            else:
                # Đây là ranh giới THỰC SỰ giữa hai ô/khối gộp khác nhau
                # (Ví dụ: ranh giới giữa A6 [cha là A3] và A7 [cha là A7])
                
                # Kiểm tra border.bottom của khối BÊN TRÊN
                cell_above = worksheet.cell(row=style_coord_above[0], column=style_coord_above[1])
                if cell_above.border.bottom and cell_above.border.bottom.style and cell_above.border.bottom.style != 'none':
                    has_bottom_border = True
                    
                # Kiểm tra border.top của khối BÊN DƯỚI
                cell_below = worksheet.cell(row=style_coord_below[0], column=style_coord_below[1])
                if cell_below.border.top and cell_below.border.top.style and cell_below.border.top.style != 'none':
                    has_top_border = True
            
            # --- (Điều kiện HOẶC) ---
            if has_bottom_border or has_top_border:
                horizontal_count += 1
        
        # --- (Logic báo cáo và ghi sổ) ---
        border_rate = horizontal_count / total_columns
        
        status = ""
        if border_rate >= border_threshold:
            last_split_row_idx = r_idx 
            status = " ✓ ỨNG VIÊN"
        
        print(f"  Ranh giới {r_idx:2d} (giữa Excel {real_row_above:2d} & {real_row_below:2d}): "
              f"{horizontal_count:2d}/{total_columns:2d} = "
              f"{border_rate:5.1%}{status}")
    
        # --- (Logic kết luận) ---
        if last_split_row_idx != -1:
            data_start_row_idx = last_split_row_idx + 1
            
            print(f"\n✓ Ranh giới CUỐI CÙNG tìm thấy tại index hàng header: {last_split_row_idx}")
            print(f"  Header: 0-{last_split_row_idx}, Data: {data_start_row_idx}+")
            return data_start_row_idx 
    
    # Thêm một cảnh báo hữu ích
    if border_threshold >= 1.0:
        print(f"\n✗ Không tìm thấy ranh giới nào >= {border_threshold*100:.0f}%.")
        print("  GỢI Ý: Threshold 100% rất nhạy cảm. Hãy thử hạ xuống 0.95 (95%).")
    else:
        print(f"\n✗ Không tìm thấy ranh giới nào >= {border_threshold*100:.0f}%.")
        
    return -1



def detect_attribute_boundary(
    header_df: pd.DataFrame, 
    data_df: pd.DataFrame 
) -> Tuple[List[int], List[int]]:
    """
    (Phiên bản V8 - Logic "Chốt" của bạn)
    Phân tích `header_df` VÀ `data_df` để tìm ranh giới Thuộc tính.
    
    Quy tắc (Heuristic) của bạn:
    Một cột là "THUỘC TÍNH" NẾU:
    1. (Logic Header) "Thân" (body) của nó trong header rỗng (do gộp dọc).
       HOẶC LÀ
    2. (Logic Data) Nó là cột kiểu "văn bản" (string/object) VÀ 
       có chứa dữ liệu gộp (phát hiện bằng cách tìm cả giá trị và NaN).
       
    Cột ĐẦU TIÊN không thỏa mãn cả 2 điều kiện trên là "vách đá" (ranh giới).
    """
    print(f"\n[detect_attribute_boundary] Phân tích {header_df.shape[1]} cột header (Logic Hybrid)...")
    
    attribute_cols_idx = []
    data_cols_idx = []
    
    total_header_rows = header_df.shape[0]
    total_cols = header_df.shape[1]

    # --- (TOÀN BỘ LOGIC BÊN DƯỚI ĐÃ ĐƯỢC VIẾT LẠI) ---
    
    # Lặp qua TẤT CẢ các cột để kiểm tra
    for c_idx in header_df.columns:
        
        # --- Check 1: Logic Header (Thân rỗng) ---
        column_body = header_df.iloc[1: , c_idx]
        body_is_empty = column_body.isna().all()
        # (Nếu thân rỗng, nó là thuộc tính)
        is_header_attr = body_is_empty
        
        # --- Check 2: Logic Data (Gộp dọc trong data) ---
        data_column = data_df.iloc[:, c_idx]
        
        # 2a. Kiểm tra kiểu dữ liệu an toàn (tránh nhầm lẫn NaN của số)
        dtype = pd_types.infer_dtype(data_column, skipna=True)
        is_string_like = dtype in ('string', 'object', 'mixed', 'unknown', 'datetime', 'date')
        
        # 2b. Kiểm tra xem có phải là cột gộp không (có cả giá trị và NaN)
        has_nans = data_column.isna().any()
        has_values = data_column.notna().any()
        
        # (Nếu là kiểu string VÀ có cả NaN/giá trị -> nó là thuộc tính gộp)
        is_data_attr = is_string_like and has_nans and has_values

        
        # --- Quyết định cuối cùng (Logic "HOẶC" của bạn) ---
        if is_header_attr or is_data_attr:
            # Nếu 1 trong 2 đúng, đây là Cột Thuộc tính
            print(f"  -> Cột {c_idx} là Cột Thuộc tính (Header: {is_header_attr}, Data: {is_data_attr})")
            attribute_cols_idx.append(c_idx)
        else:
            # Đây là "vách đá" - Cột Dữ liệu đầu tiên
            # (Nó không có thân header rỗng VÀ nó không phải là cột data gộp)
            print(f"  -> Ranh giới tại Cột {c_idx} (Không phải Thuộc tính)")
            
            # Tất cả các cột từ đây về sau ĐỀU LÀ Cột Dữ liệu
            data_cols_idx = list(range(c_idx, total_cols))
            
            # Thoát vòng lặp
            break

    print(f"\n  -> [CHỐT] Cột Thuộc tính: {attribute_cols_idx}")
    print(f"  -> [CHỐT] Cột Dữ liệu: {data_cols_idx}")
    return attribute_cols_idx, data_cols_idx

# --- [GIAI ĐOẠN 3: Trích xuất JSON] ---


def _set_nested_value(target_dict: Dict, path: List[str], value: Any):
    """
    (Hàm trợ giúp - Bánh xe)
    Đi theo `path` và gán `value` ở cấp cuối cùng.
    Ví dụ: _set_nested_value(d, ['Group 1', 'Sub 1'], 5)
    -> d['Group 1']['Sub 1'] = 5
    """
    for key in path[:-1]:
        # Nếu key chưa có, tạo 1 dict con
        target_dict = target_dict.setdefault(key, {})
    # Gán giá trị ở cấp cuối cùng
    target_dict[path[-1]] = value


def _build_header_map(header_df: pd.DataFrame, data_cols: List[int]) -> Dict[int, List[str]]:
    """
    (Hàm MỚI - Bước 2.3)
    Phân tích `header_df` và tạo "Bản đồ Header" cho các cột dữ liệu.
    
    Logic:
    1. Lấp đầy (ffill) các ô gộp (cả ngang và dọc).
    2. Đọc "dọc" từng cột để xây dựng "con đường" (path).
    
    Returns:
        Một dict (bản đồ): { column_index -> [path, to, header] }
        Ví dụ: { 5: ['(Group 1)', 'Sub-Group 1.1', 'F-Data'] }
    """
    print(f"\n[build_header_map] Đang xây dựng bản đồ cho {len(data_cols)} cột dữ liệu...")
    
    # 1. Lấp đầy (ffill) để xử lý ô gộp
    # Fill ngang (axis=1) để vá các lỗ hổng ô gộp
    header_df_filled = header_df.ffill(axis=1)
    # Fill dọc (axis=0) để lấp đầy các cấp (ví dụ: Sub-Group 1.1)
    header_df_filled = header_df_filled.ffill(axis=0)
    
    header_map = {}
    
    # Chỉ lặp qua các CỘT DỮ LIỆU
    for c_idx in data_cols:
        path = []
        last_val = None # Dùng để tránh lặp lại (ví dụ: Group 1, Group 1, Group 1...)
        
        # Lặp qua từng hàng (row_index) trong header_df
        for r_idx in header_df_filled.index:
            value = header_df_filled.loc[r_idx, c_idx]
            
            # Chỉ thêm nếu nó không NaN VÀ không bị lặp lại
            if pd.notna(value) and value != last_val:
                if isinstance(value, str):
                    path.append(value)
                else:
                    print(f"  ⚠ Cảnh báo: Giá trị không phải chuỗi ở header (hàng {r_idx}, cột {c_idx}): {value}")
                    path.append(str(value))
                last_val = value
        
        header_map[c_idx] = path
    
    # print(f"  -> Bản đồ Header (mẫu): Cột 5 -> {header_map.get(5)}")
    return header_map

def parse_table_to_long_json(
    header_df: pd.DataFrame, 
    data_df: pd.DataFrame, 
    attribute_cols: List[int], 
    data_cols: List[int]
) -> List[Dict[str, Any]]:
    """
    (Hàm MỚI - Bước 2.4 - ĐÃ SỬA LOGIC)
    Lắp ráp JSON theo định dạng "Rộng" (Wide Format)
    (Tạo MỘT object JSON cho mỗi HÀNG dữ liệu, gộp tất cả các cột).
    """
    
    final_json_list = []
    
    # --- 1. Chuẩn bị 2 "Bản đồ" ---
    
    # Bản đồ 1: "Bản đồ Header" (Tra cứu Path theo Cột)
    header_map = _build_header_map(header_df, data_cols)
    
    print("header_map",header_map)
    
    # Bản đồ 2: "Tên Thuộc tính" (Lấy tên "Ngày", "ID" từ hàng đầu)
    attribute_key_names = [header_df.iloc[0, c_idx] for c_idx in attribute_cols]
    
    print(f"[parse_table_to_long_json] Đang lấp đầy (ffill) các thuộc tính gộp...")
    filled_data_df = data_df.copy()
    filled_data_df.loc[:, attribute_cols] = filled_data_df.loc[:, attribute_cols].ffill()
    
    # --- 2. Vòng lặp Kép (ĐÃ SỬA LOGIC LẮP RÁP) ---
    
    print(f"[parse_table_to_long_json] Đang lắp ráp các HÀNG (rows)...") # Đã sửa log
    
    # Lặp qua các HÀNG DỮ LIỆU (ví dụ: index 5, 6)
    for r_idx in filled_data_df.index:
        
        # a. Lấy "Bản ghi Thuộc tính" (Attribute Record) cho hàng này
        base_record = {}
        for i, c_idx in enumerate(attribute_cols):
            key = attribute_key_names[i]
            value = filled_data_df.loc[r_idx, c_idx]
            base_record[key] = value
            # print("base_record",base_record)
        
        # --- [SỬA LỖI 1] ---
        # Tạo MỘT bản sao DUY NHẤT cho CẢ HÀNG
        # (Di chuyển ra ngoài vòng lặp 'c_idx')
        record = base_record.copy() 
        
        # b. Lặp qua các CỘT DỮ LIỆU (ví dụ: 1, 2, ..., 29)
        for c_idx in data_cols:
            
            # i. Lấy Giá trị (Value)
            value = filled_data_df.loc[r_idx, c_idx]
            
            # Bỏ qua nếu ô đó trống (không tạo JSON cho ô NaN)
            if pd.isna(value):
                continue
                
            # ii. Lấy "Con đường" (Path)
            path = header_map[c_idx]
            # print("path",path)
            
            # iii. Lắp ráp
            
            # (Dòng 'record = base_record.copy()' ĐÃ BỊ XÓA KHỎI ĐÂY)
            
            # Tạo object lồng nhau (Keys)
            nested_data_obj = {}
            _set_nested_value(nested_data_obj, path, value)
            
            # Gộp 2 phần lại
            # CẬP NHẬT (update) vào 'record' duy nhất của hàng
            record.update(nested_data_obj)
            # print("record",record)
            
            # (Dòng 'final_json_list.append(record)' ĐÃ BỊ XÓA KHỎI ĐÂY)
            
        # --- [SỬA LỖI 2] ---
        # Thêm vào kết quả cuối cùng SAU KHI lặp xong TẤT CẢ các cột
        # (Di chuyển ra ngoài vòng lặp 'c_idx')
        final_json_list.append(record)
            
    return final_json_list


def transform_all_data_in_json_to_string(
    json_data: List[Dict[str, Any]]
) -> List[Dict[str, Any]]:
    """
    (Hàm MỚI - Bước 2.5)
    Chuyển đổi tất cả giá trị trong JSON sang chuỗi (string).
    """
    
    def convert_value_to_string(value: Any) -> Any:
        if isinstance(value, dict):
            return {k: convert_value_to_string(v) for k, v in value.items()}
        elif isinstance(value, list):
            return [convert_value_to_string(v) for v in value]
        elif value is None:
            return ""
        else:
            return str(value)
    
    return [convert_value_to_string(record) for record in json_data]


if __name__ == "__main__":

# ======================================================================
# GIAI ĐOẠN TEST 1: PHÁT HIỆN BẢNG
# ======================================================================

    # # Thay đổi đường dẫn này cho đúng với file của bạn
    # # FILE_PATH = "path/to/your/image_b9d51d.xlsx"
    # FILE_PATH = "Book1.xlsx" # Giả sử file tên là report.xlsx
    # SHEET_NAME = "Sheet1"     # Thay tên sheet nếu cần

    # print(f"--- Bắt đầu phát hiện bảng trong file: {FILE_PATH} ---")
    
    # # Bạn có thể điều chỉnh 'min_width' và 'min_height'
    # table_coordinates = detect_tables(
    #     FILE_PATH, 
    #     SHEET_NAME, 
    #     min_width=2, 
    #     min_height=2
    # )
    
    # print("\n--- KẾT QUẢ CUỐI CÙNG ---")
    # if table_coordinates:
    #     for i, coords in enumerate(table_coordinates):
    #         print(f"Bảng {i+1} tìm thấy tại (1-indexed):")
    #         print(f"  - Hàng: từ {coords['min_row']} đến {coords['max_row']}")
    #         print(f"  - Cột:  từ {coords['min_col']} đến {coords['max_col']}")
    # else:
    #     print("Không tìm thấy bảng nào hợp lệ.")




# ======================================================================
# GIAI ĐOẠN TEST 2: PHÁT HIỆN BẢNG VÀ TRÍCH XUẤT DỮ LIỆU 
# ======================================================================

    # FILE_PATH = "basic_test.xlsx" # Sửa lại tên file của bạn
    # SHEET_NAME = "Sheet1"    # Sửa lại tên sheet của bạn

    # print(f"--- Bắt đầu phát hiện bảng trong file: {FILE_PATH} ---")
    
    # table_coordinates = detect_tables(
    #     FILE_PATH, 
    #     SHEET_NAME, 
    #     min_width=2,  # Giữ nguyên min_width=2, min_height=2 như bạn test
    #     min_height=2
    # )
    
    # print("\n--- KẾT QUẢ CUỐI CÙNG (PHÁT HIỆN) ---")
    # if table_coordinates:
    #     for i, coords in enumerate(table_coordinates):
    #         print(f"Bảng {i+1} tìm thấy tại (1-indexed):")
    #         print(f"  - Hàng: từ {coords['min_row']} đến {coords['max_row']}")
    #         print(f"  - Cột:  từ {coords['min_col']} đến {coords['max_col']}")
            
    #         # --- PHẦN DEBUG MỚI ---
    #         print(f"\n[DEBUG] Đang trích xuất dữ liệu thô Bảng {i+1}...")
    #         raw_data = debug_extract_data(FILE_PATH, SHEET_NAME, coords)
    #         print(type(raw_data))
    #         print(raw_data)
    #         if not raw_data.empty:
    #             print(f"--- Dữ liệu thô Bảng {i+1} (đầu & cuối): ---")
    #             # Hiển thị 5 hàng đầu và 5 hàng cuối của bảng
    #         #     with pd.option_context('display.max_rows', 10, 'display.max_columns', None):
    #         #         print(raw_data.head())
    #         # print("-" * 30)
    #         # --- KẾT THÚC PHẦN DEBUG --- 
            
    # else:
    #     print("Không tìm thấy bảng nào hợp lệ.")



# ======================================================================
# GIAI ĐOẠN TEST 3: PHÁT HIỆN HEADER VÀ TÁCH DỮ LIỆU 
# ======================================================================


    # FILE_PATH = "basic_test.xlsx" # File test của bạn
    # SHEET_NAME = "basic3"         # Sheet của bạn

    # # --- PHẢI LOAD `worksheet` TRƯỚC ---
    # try:
    #     wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    #     if SHEET_NAME not in wb.sheetnames:
    #         raise ValueError(f"Không tìm thấy sheet '{SHEET_NAME}'")
    #     worksheet = wb[SHEET_NAME]
    # except Exception as e:
    #     print(f"Lỗi khi tải workbook: {e}")
    #     exit()

    # # --- CHẠY GIAI ĐOẠN 1 (ĐỂ LẤY ĐẦU VÀO) ---
    # print(f"--- [GIAI ĐOẠN 1] Đang chạy detect_tables... ---")
    # table_coordinates = detect_tables(
    #     FILE_PATH, 
    #     SHEET_NAME, 
    #     min_width=2,
    #     min_height=2
    # )
    # print(f"--- [GIAI ĐOẠN 1] Hoàn thành: Tìm thấy {len(table_coordinates)} bảng ---")

    # # Lặp qua các bảng tìm được
    # for i, coords in enumerate(table_coordinates):
    #     print(f"\n--- Xử lý Bảng {i+1} (Hàng {coords['min_row']}->{coords['max_row']}) ---")
        
    #     raw_table_df = debug_extract_data(FILE_PATH, SHEET_NAME, coords)
        
    #     if raw_table_df.empty:
    #         continue
        
    #     # --- CHẠY HÀM MỚI (CHỈ DÙNG BORDER) ---
    #     split_point_index = detect_header_split_point(
    #         raw_table_df, 
    #         worksheet,   # Truyền worksheet
    #         coords,      # Truyền tọa độ
    #         border_threshold=0.95 # Chỉ truyền ngưỡng border
    #     )
    #     # ---
        
    #     if split_point_index != -1:
    #         # Kiểm tra trường hợp ranh giới vượt quá số hàng (hiếm gặp)
    #         if split_point_index >= len(raw_table_df.index):
    #              print(f"\n--- Kết quả Bảng {i+1}: Ranh giới ({split_point_index}) vượt quá số hàng. Bảng có thể chỉ có Header.")
    #              continue

    #         print(f"\n--- Kết quả Bảng {i+1} ---")
    #         print(f"  -> Ranh giới (Split Point) tìm thấy tại index hàng: {split_point_index}")
            
    #         header_df = raw_table_df.iloc[0 : split_point_index]
    #         data_df = raw_table_df.iloc[split_point_index : ]
            
    #         print("\n  -> [KHỐI HEADER] (Keys):")
    #         print(header_df)
    #         print("\n  -> [KHỐI DỮ LIỆU] (Values):")
    #         print(data_df)
    #         print("-" * 30)
            
    #     else:
    #         print(f"\n--- Kết quả Bảng {i+1}: Không thể xác định ranh giới Header/Data ---")

    # wb.close() # Đóng workbook sau khi xong


# ======================================================================
# GIAI ĐOẠN TEST 4: TÁCH THUỘC TÍNH TRONG HEADER VÀ IN KẾT QUẢ 
# ======================================================================

    
    # FILE_PATH = "basic_test.xlsx" # File test của bạn
    # SHEET_NAME = "basic3"         # Sheet của bạn

    # # --- PHẢI LOAD `worksheet` TRƯỚC ---
    # try:
    #     wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    #     if SHEET_NAME not in wb.sheetnames:
    #         raise ValueError(f"Không tìm thấy sheet '{SHEET_NAME}'")
    #     worksheet = wb[SHEET_NAME]
    # except Exception as e:
    #     print(f"Lỗi khi tải workbook: {e}")
    #     exit()

    # # --- CHẠY GIAI ĐOẠN 1 (ĐỂ LẤY ĐẦU VÀO) ---
    # print(f"--- [GIAI ĐOẠN 1] Đang chạy detect_tables... ---")
    # table_coordinates = detect_tables(
    #     FILE_PATH, 
    #     SHEET_NAME, 
    #     min_width=2,
    #     min_height=2
    # )
    # print(f"--- [GIAI ĐOẠN 1] Hoàn thành: Tìm thấy {len(table_coordinates)} bảng ---")

    # # Lặp qua các bảng tìm được
    # for i, coords in enumerate(table_coordinates):
    #     print(f"\n--- Xử lý Bảng {i+1} (Hàng {coords['min_row']}->{coords['max_row']}) ---")
        
    #     raw_table_df = debug_extract_data(FILE_PATH, SHEET_NAME, coords)
        
    #     if raw_table_df.empty:
    #         continue
        
    #     # --- BƯỚC 2.1: TÌM RANH GIỚI HEADER/DATA (Hàm của bạn) ---
    #     split_point_index = detect_header_split_point(
    #         raw_table_df, 
    #         worksheet,   # Truyền worksheet
    #         coords,      # Truyền tọa độ
    #         border_threshold=0.95 # Chỉ truyền ngưỡng border
    #     )
        
    #     if split_point_index != -1:
    #         if split_point_index >= len(raw_table_df.index):
    #              print(f"\n--- Kết quả Bảng {i+1}: Ranh giới ({split_point_index}) vượt quá số hàng.")
    #              continue

    #         print(f"\n--- Kết quả Bảng {i+1}: Tách Khối ---")
    #         print(f"  -> Ranh giới (Split Point) tìm thấy tại index hàng: {split_point_index}")
            
    #         header_df = raw_table_df.iloc[0 : split_point_index]
    #         data_df = raw_table_df.iloc[split_point_index : ]
            
    #         print("\n  -> [KHỐI HEADER] (Keys):")
    #         print(header_df.head()) # In 5 dòng đầu
            
    #         # --- BƯỚC 2.2: TÌM RANH GIỚI THUỘC TÍNH (Hàm MỚI) ---
    #         attribute_cols, data_cols = detect_attribute_boundary(header_df)
            
    #         print("-" * 30)
            
    #     else:
    #         print(f"\n--- Kết quả Bảng {i+1}: Không thể xác định ranh giới Header/Data ---")

    # wb.close() # Đóng workbook sau khi xong

# ======================================================================
# GIAI ĐOẠN TEST 5: CHẠY TOÀN BỘ VÀ IN KẾT QUẢ JSON
# ======================================================================

    

    FILE_PATH = "Book3.xlsx" 
    # FILE_PATH = "cloned_excel.xlsx" 
    # SHEET_NAME = "Sheet1" 
    SHEET_NAME = "Sheet1" 

    # --- PHẢI LOAD `worksheet` TRƯỚC ---
    try:
        wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
        if SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"Không tìm thấy sheet '{SHEET_NAME}'")
        worksheet = wb[SHEET_NAME]
        ## transform to string
        
        
        
        # Tạo merged_map MỘT LẦN ở đây
        merged_map = _create_merged_cell_map(worksheet) 
        print(f"Đã tạo bản đồ ô gộp. (Phát hiện {len(merged_map)} ô con)")
        
    except Exception as e:
        print(f"Lỗi khi tải workbook: {e}")
        exit()

    # --- CHẠY GIAI ĐOẠN 1 (ĐỂ LẤY ĐẦU VÀO) ---
    print(f"\n--- [GIAI ĐOẠN 1] Đang chạy detect_tables... ---")
    table_coordinates = detect_tables(
        FILE_PATH, 
        SHEET_NAME, 
        min_width=2,
        min_height=2
    )
    
    print("table_coordinates",table_coordinates)
    print(f"--- [GIAI ĐOẠN 1] Hoàn thành: Tìm thấy {len(table_coordinates)} bảng ---")

    all_parsed_data = [] 

    # Lặp qua các bảng tìm được
    for i, coords in enumerate(table_coordinates):
        print(f"\n--- Xử lý Bảng {i+1} (Hàng {coords['min_row']}->{coords['max_row']}) ---")
        
        raw_table_df = debug_extract_data(FILE_PATH, SHEET_NAME, coords)
        # raw_table_df = transform_dataframe_to_all_string(raw_table_df)

        
        if raw_table_df.empty:
            continue
        
        # --- BƯỚC 2.1: GỌI HÀM ĐÃ SỬA LỖI ---
        # (Sử dụng tên hàm `detect_header_split_point` như bạn gọi)
        split_point_index = detect_header_split_point(
            raw_table_df, 
            worksheet,
            coords,
            merged_map,    # Truyền map vào
            border_threshold=0.98 # <-- THAY ĐỔI QUAN TRỌNG: 1.0 -> 0.95
        )
        
        if split_point_index != -1:
            if split_point_index >= len(raw_table_df.index):
                 print(f"\n--- Kết quả Bảng {i+1}: Ranh giới ({split_point_index}) vượt quá số hàng.")
                 continue

            header_df = raw_table_df.iloc[0 : split_point_index]
            # print("raw_table_df",raw_table_df.head(5))
            # print(header_df.head())
            ## transform header to string
            data_df = raw_table_df.iloc[split_point_index : ]
            
            # --- BƯỚC 2.2: TÌM RANH GIỚI THUỘC TÍNH ---
            attribute_cols, data_cols = detect_attribute_boundary(header_df, data_df)
            print(attribute_cols, data_cols)
            # --- BƯỚC 2.3 & 2.4: LẮP RÁP JSON ---
            try:
                # Chạy hàm parse JSON (Định dạng "Dài")
                json_output = parse_table_to_long_json(
                    header_df, 
                    data_df, 
                    attribute_cols, 
                    data_cols
                )

            
                all_parsed_data.extend(json_output)
                print(f"\n--- [GIAI ĐOẠN 2] Parse Bảng {i+1} thành công. Tạo ra {len(json_output)} bản ghi JSON.")

            except Exception as e:
                print(f"LỖI khi parse Bảng {i+1}: {e}")
                import traceback
                traceback.print_exc()

            print("-" * 30)
            
        else:
            print(f"\n--- Kết quả Bảng {i+1}: Không thể xác định ranh giới Header/Data ---")

    wb.close() # Đóng workbook sau khi xong
    
    print("\n--- [HOÀN THÀNH] Đã xử lý tất cả các bảng. ---")
    
    # In toàn bộ kết quả cuối cùng
    print("\n--- TỔNG KẾT JSON ---")
    
    # from pprint import pprint
    
    # all_parsed_data = transform_all_data_in_json_to_string(all_parsed_data)
    # pprint(all_parsed_data)

    # from helper.json_helper import convert_numpy_types
    # converted_data = convert_numpy_types(all_parsed_data)
    json_response = json.dumps(all_parsed_data, indent=2, ensure_ascii=False)

    ## Save to file
    OUTPUT_FILE = "final_output_test.json"
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write(json_response)
    print(f"✅ Đã lưu kết quả JSON vào: {OUTPUT_FILE}")