import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

# --- 1. Khởi tạo Workbook và Worksheet ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"

# --- 2. Định nghĩa các Style (Kiểu định dạng) ---

# Màu nền xám (giống trong ảnh)
grey_fill = PatternFill(start_color="D9D9D9",
                       end_color="D9D9D9",
                       fill_type="solid")

# Đường viền mỏng cho tất cả các ô
thin_side = Side(style='thin')
thin_border = Border(left=thin_side, 
                     right=thin_side, 
                     top=thin_side, 
                     bottom=thin_side)

# Căn giữa nội dung
center_align = Alignment(horizontal='center', 
                         vertical='center', 
                         wrap_text=True)

# --- 3. Thiết lập độ rộng cột (ước chừng) ---
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 12
# Thiết lập độ rộng cho các cột data (C đến V)
for col_letter in [chr(i) for i in range(ord('C'), ord('V') + 1)]:
    ws.column_dimensions[col_letter].width = 6

# --- 4. Tạo Tiêu đề (Header) - Hàng 4 và 5 ---

# Ô gộp A4:A5
ws['A4'] = "責任区分"
ws.merge_cells('A4:A5')

# Ô gộp B4:B5
ws['B4'] = "裏面"
ws.merge_cells('B4:B5')

# Tạo tiêu đề 1H, 2H, ..., 10H và các cột con
for i in range(10):
    hour_col_start_num = 3 + (i * 2) # C, E, G, ... (3, 5, 7, ...)
    hour_col_end_num = hour_col_start_num + 1
    
    # Gộp ô cho "1H", "2H", ...
    header_cell = ws.cell(row=4, column=hour_col_start_num)
    header_cell.value = f"{i+1}H"
    ws.merge_cells(start_row=4, start_column=hour_col_start_num,
                   end_row=4, end_column=hour_col_end_num)
    
    # Đặt giá trị cho "外板" và "内板"
    ws.cell(row=5, column=hour_col_start_num).value = "外板"
    ws.cell(row=5, column=hour_col_end_num).value = "内板"

# --- 5. Tạo Tiêu đề Hàng (Cột A và B) ---
row_headers_b = [
    "ブツ", "黒色ブ", "カス", "黒色カ", "ヤニ", "ブツタ", "カスタ", 
    "ハジキ", "糸", "毛", "ホコリ", "ハジキ", "汚れ", "カブリ", 
    "水滴", "サワリ", "干渉キ"
]

# Điền dữ liệu cho cột B từ hàng 6
for i, header in enumerate(row_headers_b):
    ws.cell(row=6 + i, column=2).value = header

# Gộp ô cho "上塗り責任" ở cột A
ws['A17'] = "上塗り責任"
ws.merge_cells('A17:A22') # Gộp từ hàng 17 (ハジキ) đến 22 (干渉キ)

# --- 6. Điền dữ liệu vào các ô ---
# (Hàng, Cột) - Lưu ý: C='C', D='D', ...
data = {
    (7, 'C'): 2,   # 黒色ブ, 1H-外板
    (8, 'D'): 1,   # カス, 1H-内板
    (8, 'E'): 2,   # カス, 2H-外板
    (8, 'F'): 1,   # カス, 2H-内板
    (9, 'F'): 1,   # 黒色カ, 2H-内板
    (8, 'G'): 2,   # カス, 3H-外板
    (8, 'H'): 1,   # カス, 3H-内板
    (14, 'H'): 1,  # 糸, 3H-内板
    (8, 'I'): 2,   # カス, 4H-外板
    (8, 'J'): 1,   # カス, 4H-内板
    (8, 'K'): 2,   # カス, 5H-外板
    (8, 'L'): 2,   # カス, 5H-内板
    (8, 'M'): 2,   # カス, 6H-外板
    (7, 'N'): 1,   # 黒色ブ, 6H-内板
    (8, 'O'): 2,   # カス, 7H-外板
    (14, 'P'): 1,  # 糸, 7H-内板
    (7, 'Q'): 1,   # 黒色ブ, 8H-外板
    (14, 'R'): 1,  # 糸, 8H-内板
    (7, 'S'): 8,   # 黒色ブ, 9H-外板
    (17, 'S'): 1,  # ハジキ, 9H-外板
    (7, 'T'): 1,   # 黒色ブ, 9H-内板
    (8, 'V'): 1,   # カス, 10H-内板
    (16, 'V'): 1,  # ホコリ, 10H-内板
}

for (row, col_char), value in data.items():
    ws[f"{col_char}{row}"] = value

# --- 7. Áp dụng Style (Fill, Border, Alignment) ---
# Lặp qua tất cả các ô trong phạm vi
for row in ws.iter_rows(min_row=4, max_row=22, min_col=1, max_col=22): # 22 là cột 'V'
    for cell in row:
        # Áp dụng đường viền và căn giữa cho tất cả
        cell.border = thin_border
        cell.alignment = center_align
        
        # Áp dụng màu nền xám cho vùng dữ liệu (từ cột C và từ hàng 6)
        if cell.column >= 3 and cell.row >= 6:
            cell.fill = grey_fill

# Ghi đè lại màu nền cho các ô có dữ liệu (để chúng không bị xám)
for (row, col_char) in data.keys():
    ws[f"{col_char}{row}"].fill = PatternFill(fill_type=None) # Xóa fill

# --- 8. Lưu file ---
file_name = "cloned_excel.xlsx"
wb.save(file_name)

print(f"Đã tạo file '{file_name}' thành công!")