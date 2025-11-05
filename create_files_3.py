import openpyxl
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)
from openpyxl.utils import get_column_letter

def create_management_board(filename="TienDoQuanLy.xlsx"):
    """
    Tạo một file Excel mô phỏng bảng quản lý tiến độ từ hình ảnh.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Management Board"

    # --- Định nghĩa các style ---
    
    # Màu nền
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    cyan_fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid") # Gần giống màu vàng chanh

    # Đường viền
    thin_side = Side(style='thin')
    thick_side = Side(style='thick')
    green_side = Side(style='thick', color='008000')

    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    thick_border_outline = Border(left=thick_side, right=thick_side, top=thick_side, bottom=thick_side)
    
    # Căn lề
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    vertical_text_align = Alignment(horizontal='center', vertical='center', text_rotation=255)
    
    # Phông chữ
    title_font = Font(name='Meiryo UI', size=18, bold=True)
    header_font = Font(name='Meiryo UI', size=9, bold=True)
    bold_font = Font(name='Meiryo UI', size=11, bold=True)
    red_font = Font(name='Meiryo UI', size=9, color="FF0000")
    dot_font = Font(name='Meiryo UI', size=18, bold=True) # Dấu chấm to

    # --- Thiết lập Chiều rộng Cột và Chiều cao Hàng (Ước lượng) ---
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 4
    ws.column_dimensions['C'].width = 6
    for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        ws.column_dimensions[col].width = 5
        
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[3].height = 25
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 40
    ws.row_dimensions[6].height = 20
    ws.row_dimensions[7].height = 20
    ws.row_dimensions[8].height = 25
    ws.row_dimensions[9].height = 25

    # --- Hàng 1: Tiêu đề ---
    ws.merge_cells('A1:AG1')
    cell = ws['A1']
    cell.value = "塗完P工程可動管理板"
    cell.font = title_font
    cell.alignment = Alignment(horizontal='left', vertical='center', indent=1) # Tiêu đề căn trái

    # --- Hàng 3: Thông tin ngày tháng ---
    ws['A3'].value = "11月5日"
    ws['A3'].font = Font(name='Meiryo UI', size=14, bold=True)
    ws['C3'].value = "水曜日"
    ws['C3'].font = bold_font
    ws['E3'].value = "1直"
    ws['E3'].font = bold_font
    ws['F3'].value = "A / B"
    ws['F3'].font = bold_font
    
    ws['O3'].value = "○ A"
    ws['P3'].value = "98%"
    ws['W3'].value = "○ T"
    ws['X3'].value = "98%可動"
    
    # Đường kẻ đen dày dưới hàng 3
    for col in range(2, 34): # Cột B đến AG
        ws.cell(row=3, column=col).border = Border(bottom=thick_side)

    # --- Hàng 4-7: Khối Header phức tạp ---
    
    # Khối viền xanh (B4:B7)
    for row in range(4, 8):
        ws.cell(row=row, column=2).border = Border(left=green_side, right=green_side, top=green_side, bottom=green_side)

    # Cột 'P工程' (C4:C7)
    ws.merge_cells('C4:C7')
    cell = ws['C4']
    cell.value = "P工程"
    cell.alignment = center_align
    cell.font = header_font
    cell.border = thick_border_outline

    # Khối 'ライン' (D4:H7)
    ws.merge_cells('D4:H4')
    cell = ws['D4']
    cell.value = "ライン"
    cell.alignment = center_align
    cell.font = header_font
    cell.fill = grey_fill
    
    ws.merge_cells('D5:D7')
    cell = ws['D5']
    cell.value = "対応"
    cell.alignment = vertical_text_align
    cell.font = header_font
    cell.fill = grey_fill

    ws.merge_cells('E5:E7')
    cell = ws['E5']
    cell.value = "X/W"
    cell.alignment = vertical_text_align
    cell.font = header_font
    cell.fill = grey_fill

    ws.merge_cells('F5:F7')
    cell = ws['F5']
    cell.value = "X/W"
    cell.alignment = vertical_text_align
    cell.font = header_font
    cell.fill = grey_fill

    ws.merge_cells('G5:H5')
    cell = ws['G5']
    cell.value = "ライン停止" # (Đoán chữ)
    cell.alignment = center_align
    cell.font = header_font
    cell.fill = grey_fill
    
    ws['G6'].value = "CAC"
    ws['H6'].value = "F/C"
    ws['G7'].value = "*"
    ws['H7'].value = "*"
    
    # Áp dụng style cho các ô con
    for r in range(4, 8):
        for c in range(4, 9):
            cell = ws.cell(row=r, column=c)
            if not cell.value: # Áp dụng fill cho cả ô đã merge
                cell.fill = grey_fill
            if r >= 6 and c >= 7:
                cell.alignment = center_align
                cell.font = Font(name='Meiryo UI', size=14, bold=True)
            cell.border = thin_border
            
    # Áp dụng viền dày cho khối D4:H7
    for r in range(4, 8):
        ws.cell(row=r, column=4).border = Border(left=thick_side, top=thin_side, bottom=thin_side, right=thin_side)
        ws.cell(row=r, column=8).border = Border(left=thin_side, top=thin_side, bottom=thin_side, right=thick_side)
    for c in range(4, 9):
        ws.cell(row=4, column=c).border = Border(left=thin_side, top=thick_side, bottom=thin_side, right=thin_side)
        ws.cell(row=7, column=c).border = Border(left=thin_side, top=thin_side, bottom=thick_side, right=thin_side)
    # Góc
    ws['D4'].border = Border(left=thick_side, top=thick_side)
    ws['H4'].border = Border(right=thick_side, top=thick_side)
    ws['D7'].border = Border(left=thick_side, bottom=thick_side)
    ws['H7'].border = Border(right=thick_side, bottom=thick_side)


    # Cột '設備' (I4:I7)
    ws.merge_cells('I4:I7')
    cell = ws['I4']
    cell.value = "設備"
    cell.alignment = center_align
    cell.font = header_font
    cell.fill = grey_fill
    cell.border = thick_border_outline

    # Khối tiêu đề dữ liệu (J4:P7)
    headers_row4 = ["カ", "ラ", "停止", "要求", "・", "・", "・"]
    headers_row5 = ["", "", "分", "事項", "・", "・", "・"]
    data_row6 = ["*", "1", 28, 4, "", "", ""]
    data_row7 = ["*", "...", 28, 4, "", "", ""] # '...' thay cho biểu tượng
    
    for i, val in enumerate(headers_row4, 10):
        cell = ws.cell(row=4, column=i)
        cell.value = val
        cell.alignment = center_align
        cell.font = header_font
        cell.border = thin_border
        
    for i, val in enumerate(headers_row5, 10):
        cell = ws.cell(row=5, column=i)
        cell.value = val
        cell.alignment = center_align
        cell.font = header_font
        cell.border = thin_border
        
    ws['J5'].value = "タクト"
    ws['J5'].font = red_font
    ws['K5'].value = "見込み"
    ws['K5'].font = red_font
    
    ws.merge_cells('L6:L7')
    ws.merge_cells('M6:M7')

    for i, val in enumerate(data_row6, 10):
        cell = ws.cell(row=6, column=i)
        cell.value = val
        cell.alignment = center_align
        cell.font = header_font
        cell.border = thin_border

    for i, val in enumerate(data_row7, 10):
            cell = ws.cell(row=7, column=i)
            
            # Chỉ gán giá trị nếu cột không phải là L (12) hoặc M (13)
            # vì giá trị đã được gán ở ô L6/M6 (ô đã merge)
            if i not in [12, 13]: 
                cell.value = val
                
            # Áp dụng style cho tất cả các ô, kể cả ô đã merge
            cell.alignment = center_align
            cell.font = header_font
            cell.border = thin_border
        
    # Khối 'アルウェット' (Q4:Y7)
    ws.merge_cells('Q4:Y4')
    ws['Q4'].value = "アルウェット"
    ws.merge_cells('Q5:Y5')
    ws['Q5'].value = "塗装設備全停止"
    ws.merge_cells('Q6:S6')
    ws['Q6'].value = "第1上塗"
    ws.merge_cells('T6:V6')
    ws['T6'].value = "第2上塗"
    ws.merge_cells('W6:Y6')
    ws['W6'].value = "C#F" # (Đoán chữ)
    
    for r in range(4, 8):
        for c in range(17, 26): # Q đến Y
            cell = ws.cell(row=r, column=c)
            cell.fill = grey_fill
            cell.alignment = center_align
            cell.font = header_font
            cell.border = thin_border
    
    for c in range(17, 26): # Q đến Y
        for r in [7]:
            ws.cell(row=r, column=c).value = "・"
            ws.cell(row=r, column=c).font = dot_font

    # Khối 'ハイエース' (Z4:AC7)
    ws.merge_cells('Z4:AC4')
    ws['Z4'].value = "ハイエース"
    for r in range(4, 8):
        for c in range(26, 30): # Z đến AC
            cell = ws.cell(row=r, column=c)
            cell.fill = grey_fill
            cell.alignment = center_align
            cell.font = header_font
            cell.border = thin_border
    for r in range(5, 8):
        for c in range(26, 30): # Z đến AC
            ws.cell(row=r, column=c).value = "・"
            ws.cell(row=r, column=c).font = dot_font

    # Khối 'コート' (AD4:AE7)
    ws.merge_cells('AD4:AE4')
    ws['AD4'].value = "コート"
    for r in range(4, 8):
        for c in range(30, 32): # AD đến AE
            cell = ws.cell(row=r, column=c)
            cell.fill = grey_fill
            cell.alignment = center_align
            cell.font = header_font
            cell.border = thin_border
    for r in range(5, 8):
        for c in range(30, 32): # AD đến AE
            ws.cell(row=r, column=c).value = "・"
            ws.cell(row=r, column=c).font = dot_font
            
    # Viền dày cho các khối chính
    for r in range(4, 8):
        ws.cell(row=r, column=17).border = Border(left=thick_side, top=thin_side, bottom=thin_side, right=thin_side)
        ws.cell(row=r, column=25).border = Border(left=thin_side, top=thin_side, bottom=thin_side, right=thick_side)
        ws.cell(row=r, column=26).border = Border(left=thick_side, top=thin_side, bottom=thin_side, right=thin_side)
        ws.cell(row=r, column=29).border = Border(left=thin_side, top=thin_side, bottom=thin_side, right=thick_side)
        ws.cell(row=r, column=30).border = Border(left=thick_side, top=thin_side, bottom=thin_side, right=thin_side)
        ws.cell(row=r, column=31).border = Border(left=thin_side, top=thin_side, bottom=thin_side, right=thick_side)

    for c in range(17, 32):
        ws.cell(row=4, column=c).border = Border(left=thin_side, top=thick_side, bottom=thin_side, right=thin_side)
        ws.cell(row=7, column=c).border = Border(left=thin_side, top=thin_side, bottom=thick_side, right=thin_side)
    
    # Góc
    ws['Q4'].border = Border(left=thick_side, top=thick_side)
    ws['Y4'].border = Border(right=thick_side, top=thick_side)
    ws['Q7'].border = Border(left=thick_side, bottom=thick_side)
    ws['Y7'].border = Border(right=thick_side, bottom=thick_side)
    #... (tương tự cho các khối khác)


    # --- Hàng 8 & 9: Dữ liệu màu ---
    for c in range(3, 34): # Cột C đến AG
        ws.cell(row=8, column=c).fill = cyan_fill
        ws.cell(row=8, column=c).value = "..." # Placeholder
        ws.cell(row=8, column=c).alignment = center_align
        ws.cell(row=9, column=c).fill = yellow_fill
        ws.cell(row=9, column=c).value = "..." # Placeholder
        ws.cell(row=9, column=c).alignment = center_align

    # --- Lưu file ---
    wb.save(filename)
    print(f"Đã tạo file '{filename}' thành công!")

if __name__ == "__main__":
    create_management_board()