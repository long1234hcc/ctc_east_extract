import pandas as pd
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from collections import deque
from typing import List, Dict, Set, Tuple, Any
import numpy as np
import json


# ======================================================================
# GIAI ĐOẠN 1: PHÁT HIỆN BẢNG (Giữ nguyên từ trước)
# ======================================================================

# --- BƯỚC 1: TẠO BẢN ĐỒ TRA CỨU Ô GỘP ---
def _create_merged_cell_map(ws: Worksheet) -> Dict[Tuple[int, int], Tuple[int, int]]:
    """
    Tạo một dict (map) để tra cứu ô "cha" (ô top-left chứa style)
    từ bất kỳ tọa độ ô "con" nào.
    
    Returns:
        Dict[(con_r, con_c), (cha_r, cha_c)]
    """
    merged_map = {}
    # Lặp qua tất cả các dải ô gộp trong sheet
    for merged_range in ws.merged_cells.ranges:
        # Lấy tọa độ (1-based index) của dải ô
        min_col, min_row, max_col, max_row = merged_range.bounds
        
        # Tọa độ ô "cha" (ô top-left)
        parent_coord = (min_row, min_col)
        
        # Lặp qua tất cả ô "con" trong dải (bao gồm cả ô cha)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                # Ánh xạ ô con về ô cha
                merged_map[(r, c)] = parent_coord
    return merged_map

# --- BƯỚC 2: TẠO BẢN ĐỒ NHIỆT BORDER (ĐÃ XỬ LÝ Ô GỘP) ---
def _create_border_heatmap(ws: Worksheet, merged_map: Dict) -> List[List[bool]]:
    """
    Tạo một bản đồ 2D (heatmap) của sheet.
    True = "Đất" (Ô này có border, hoặc là 1 phần của ô gộp có border)
    False = "Biển" (Ô này không có border)
    
    Bản đồ này sử dụng 0-based index để dễ dàng cho Bước 3.
    """
    max_r, max_c = ws.max_row, ws.max_column
    
    # Tạo bản đồ rỗng (0-indexed)
    # heatmap[hàng][cột]
    heatmap = [[False for _ in range(max_c)] for _ in range(max_r)]

    # Lặp qua từng ô trong sheet (1-indexed)
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            
            # 1. Tìm tọa độ ô chứa style
            # Mặc định là chính nó
            style_coord = (r, c)
            if (r, c) in merged_map:
                # Nếu là ô con, lấy tọa độ ô cha
                style_coord = merged_map[(r, c)]
            
            # 2. Lấy ô style từ tọa độ
            style_cell = ws.cell(row=style_coord[0], column=style_coord[1])
            
            # 3. Kiểm tra border của ô style
            b = style_cell.border
            # Chỉ cần 1 cạnh có style là coi như ô đó có border
            if (b.left.style or b.right.style or b.top.style or b.bottom.style):
                # Đánh dấu "Đất" (True) vào heatmap (0-indexed)
                heatmap[r-1][c-1] = True
                
    return heatmap

# --- BƯỚC 3: TÌM "CỤM BORDER" (BFS) ---
def _find_clusters(heatmap: List[List[bool]]) -> List[List[Tuple[int, int]]]:
    """
    Chạy thuật toán BFS (Breadth-First Search) trên heatmap
    để tìm các "quần đảo" (cụm) các ô "Đất" (True) liền kề nhau.
    
    Returns:
        List các cụm, mỗi cụm là 1 List các tọa độ (r, c) (0-indexed).
    """
    if not heatmap: 
        return []
        
    rows = len(heatmap)
    cols = len(heatmap[0])
    
    visited = set()  # Set chứa các tọa độ (r, c) (0-indexed) đã ghé thăm
    clusters = []    # List chứa các cụm

    for r in range(rows):
        for c in range(cols):
            # Nếu ô này là "Đất" (True) và chưa được ghé thăm
            if heatmap[r][c] and (r, c) not in visited:
                
                # Bắt đầu một cụm mới
                new_cluster = []
                q = deque([(r, c)]) # Hàng đợi cho BFS
                visited.add((r, c))

                while q:
                    curr_r, curr_c = q.popleft()
                    # Thêm tọa độ (0-indexed) vào cụm
                    new_cluster.append((curr_r, curr_c))

                    # Kiểm tra 4 hướng lân cận (trên, dưới, trái, phải)
                    for dr, dc in [(0, 1), (0, -1), (1, 0), (-1, 0)]:
                        nr, nc = curr_r + dr, curr_c + dc

                        # Kiểm tra xem có nằm trong ranh giới bản đồ không
                        if 0 <= nr < rows and 0 <= nc < cols:
                            # Nếu ô lân cận là "Đất" và chưa ghé thăm
                            if heatmap[nr][nc] and (nr, nc) not in visited:
                                visited.add((nr, nc))
                                q.append((nr, nc))
                
                # Sau khi while kết thúc, thêm cụm mới vào danh sách
                clusters.append(new_cluster)
                
    return clusters

# --- BƯỚC 4: LỌC CỤM & LẤY TỌA ĐỘ (BOUNDING BOX) ---
def _filter_and_get_boundaries(clusters: List[List[Tuple[int, int]]], 
                               min_width: int = 5, 
                               min_height: int = 3) -> List[Dict[str, int]]:
    """
    Lặp qua các cụm, lọc bỏ "nhiễu" (cụm quá nhỏ),
    và trả về tọa độ (bounding box) của các "bảng" hợp lệ.
    
    Tọa độ trả về là 1-indexed (để khớp với Excel).
    """
    final_table_boundaries = []
    
    for cluster in clusters:
        if not cluster:
            continue
            
        # Lấy tất cả tọa độ (0-indexed) của cụm
        all_r = [r for r, c in cluster]
        all_c = [c for r, c in cluster]
        
        # Tìm min/max (0-indexed)
        min_r, max_r = min(all_r), max(all_r)
        min_c, max_c = min(all_c), max(all_c)
        
        # Tính toán kích thước
        width = max_c - min_c + 1
        height = max_r - min_r + 1
        
        # Áp dụng bộ lọc (heuristic)
        if width >= min_width and height >= min_height:
            
            # Nếu đủ lớn, lưu lại tọa độ (chuyển về 1-indexed)
            final_table_boundaries.append({
                'min_row': min_r + 1,
                'max_row': max_r + 1,
                'min_col': min_c + 1,
                'max_col': max_c + 1
            })
            
    return final_table_boundaries

# --- HÀM TỔNG HỢP (MAIN FUNCTION) ---
def detect_tables(file_path: str, sheet_name: str, 
                  min_width: int = 5, 
                  min_height: int = 3) -> List[Dict[str, int]]:
    """
    Phát hiện tất cả các "bảng" (được định nghĩa bằng border)
    trong một sheet Excel.
    
    Args:
        file_path: Đường dẫn đến file Excel.
        sheet_name: Tên sheet cần xử lý.
        min_width: Chiều rộng tối thiểu để coi là 1 bảng (ý tưởng "line > 5").
        min_height: Chiều cao tối thiểu để coi là 1 bảng.
        
    Returns:
        Một list các dict, mỗi dict chứa tọa độ 1-indexed của bảng.
        Ví dụ: [{'min_row': 2, 'max_row': 12, 'min_col': 1, 'max_col': 26}]
    """
    try:
        # data_only=True để đọc giá trị (nếu cần), không phải công thức
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            print(f"Lỗi: Không tìm thấy sheet '{sheet_name}' trong file.")
            return []
        ws = wb[sheet_name]
    except Exception as e:
        print(f"Lỗi khi tải file hoặc sheet: {e}")
        return []

    # --- Chạy 4 bước của Giai đoạn 1 ---
    
    # Bước 1:
    print(f"Bước 1: Đang tạo bản đồ ô gộp...")
    merged_map = _create_merged_cell_map(ws)
    print(f"Bước 1: Hoàn thành. Tìm thấy {len(merged_map)} ô con trong các ô gộp.")
    
    # Bước 2:
    print(f"Bước 2: Đang tạo bản đồ nhiệt border (có xử lý ô gộp)...")
    heatmap = _create_border_heatmap(ws, merged_map)
    print("Bước 2: Hoàn thành.")
    
    # Bước 3:
    print(f"Bước 3: Đang tìm các cụm border...")
    clusters = _find_clusters(heatmap)
    print(f"Bước 3: Hoàn thành. Tìm thấy {len(clusters)} cụm.")
    
    # Bước 4:
    print(f"Bước 4: Đang lọc cụm và lấy tọa độ (min_width={min_width}, min_height={min_height})...")
    boundaries = _filter_and_get_boundaries(clusters, min_width, min_height)
    print(f"Bước 4: Hoàn thành. Tìm thấy {len(boundaries)} bảng hợp lệ.")
    
    wb.close()
    return boundaries







def debug_extract_data(file_path: str, sheet_name: str, 
                       boundary: Dict[str, int]) -> pd.DataFrame:
    """
    Đọc và trả về dữ liệu thô (raw data) từ BÊN TRONG một tọa độ (boundary)
    đã được phát hiện, dùng cho mục đích kiểm tra (debug).
    
    Tọa độ boundary nhận vào là 1-indexed.
    """
    
    # 1. Chuyển đổi tọa độ 1-indexed (từ detect_tables) 
    #    sang 0-indexed (cho pandas)
    
    # Hàng 3 (1-indexed) -> skiprows=2 (bỏ qua hàng 0, 1)
    skip_rows = boundary['min_row'] - 1
    
    # Số hàng cần đọc
    num_rows = boundary['max_row'] - boundary['min_row'] + 1
    
    # Cột 1 (1-indexed) -> cột 0 (0-indexed)
    # Cột 26 (1-indexed) -> cột 25 (0-indexed)
    # Chúng ta cần list [0, 1, ..., 25]
    cols_to_use = list(range(
        boundary['min_col'] - 1,  # (1-1) = 0
        boundary['max_col']       # (26) -> range() sẽ dừng ở 25
    ))
    
    if not cols_to_use:
        print("Lỗi: Không có cột nào để đọc.")
        return pd.DataFrame()

    # 2. Đọc file Excel chỉ trong phạm vi đã định
    try:
        raw_table_df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=None,        # Không giả định header, đọc thô
            skiprows=skip_rows,   # Bỏ qua các hàng bên trên
            nrows=num_rows,     # Chỉ đọc số hàng của bảng
            usecols=cols_to_use   # Chỉ đọc các cột của bảng
        )
        
        # Đặt lại index cột để dễ nhìn (0, 1, 2...)
        raw_table_df.columns = range(raw_table_df.shape[1])
        
        return raw_table_df
        
    except Exception as e:
        print(f"Lỗi khi trích xuất dữ liệu debug: {e}")
        return pd.DataFrame()
    

# ======================================================================
# GIAI ĐOẠN 2: TRÍCH XUẤT JSON (Code mới)
# ======================================================================
import pandas as pd
import json
from typing import Dict, List, Any, Optional, Tuple
import re


class DynamicExcelParser:
    """
    Parser động cho bảng Excel với header nhiều cấp.
    Tự động phát hiện cấu trúc và chuyển đổi sang nested JSON.
    """
    
    def __init__(self, df: pd.DataFrame):
        self.df = df
        self.header_end_row = 0
        self.data_start_row = 0
        self.column_structure = []
        
    def parse(self) -> Dict[str, Any]:
        """Parse toàn bộ DataFrame sang nested JSON."""
        
        # Bước 1: Tìm ranh giới giữa header và data
        self._detect_header_boundary()
        
        # Bước 2: Parse cấu trúc header
        self._parse_header_structure()
        
        # Bước 3: Parse dữ liệu
        data_rows = self._parse_data_rows()
        
        return {
            "metadata": {
                "header_rows": self.header_end_row,
                "data_start_row": self.data_start_row,
                "total_columns": len(self.column_structure),
                "column_structure": self.column_structure
            },
            "data": data_rows
        }
    
    def _detect_header_boundary(self):
        """
        Tự động phát hiện hàng nào là ranh giới giữa header và data.
        Sử dụng heuristic: hàng đầu tiên có pattern như ID-1, ID-2, hoặc ngày tháng thực.
        """
        
        for idx in range(len(self.df)):
            row = self.df.iloc[idx]
            
            # Kiểm tra cột thứ 2 (thường là ID)
            if pd.notna(row[1]):
                val = str(row[1]).strip()
                
                # Pattern: ID-số hoặc số thuần túy (không phải text mô tả)
                if re.match(r'^ID-?\d+$', val, re.IGNORECASE) or \
                   (val.isdigit() and int(val) < 1000):  # ID dạng số nhỏ
                    self.data_start_row = idx
                    self.header_end_row = idx
                    break
            
            # Nếu có nhiều ô liên tiếp chứa số (dữ liệu thực)
            numeric_count = sum(1 for v in row[2:] if self._is_numeric(v))
            if numeric_count > len(row) * 0.3:  # >30% là số
                self.data_start_row = idx
                self.header_end_row = idx
                break
        
        if self.header_end_row == 0:
            # Fallback: giả sử 5 hàng đầu là header
            self.header_end_row = min(5, len(self.df) - 1)
            self.data_start_row = self.header_end_row
    
    def _is_numeric(self, val) -> bool:
        """Kiểm tra giá trị có phải số không."""
        if pd.isna(val):
            return False
        try:
            float(val)
            return True
        except:
            return False
    
    def _parse_header_structure(self):
        """
        Parse cấu trúc header động, tự động phát hiện các nhóm và nhóm con.
        """
        
        header_rows = []
        for idx in range(self.header_end_row):
            header_rows.append(self.df.iloc[idx].values.tolist())
        
        if not header_rows:
            # Không có header, mỗi cột là một field đơn giản
            self.column_structure = [
                {"col_index": i, "path": [f"Column_{i}"], "name": f"Column_{i}"}
                for i in range(len(self.df.columns))
            ]
            return
        
        # Parse từng cột
        num_cols = len(self.df.columns)
        
        for col_idx in range(num_cols):
            col_path = self._build_column_path(header_rows, col_idx)
            
            self.column_structure.append({
                "col_index": col_idx,
                "path": col_path,
                "name": col_path[-1] if col_path else f"Column_{col_idx}",
                "full_path": " > ".join(col_path)
            })
    
    def _build_column_path(self, header_rows: List[List], col_idx: int) -> List[str]:
        """
        Xây dựng path phân cấp cho một cột từ các hàng header.
        
        Logic:
        - Đọc từ trên xuống dưới
        - Bỏ qua NaN
        - Phát hiện merged cells (giá trị trải dài nhiều cột)
        - Xây dựng path: [Group] -> [SubGroup] -> [Column Name]
        """
        
        path = []
        
        for row_idx, row in enumerate(header_rows):
            val = row[col_idx]
            
            # Bỏ qua NaN
            if pd.isna(val):
                # Kiểm tra xem có phải merged cell không (tìm giá trị gần nhất bên trái)
                merged_val = self._find_merged_value(row, col_idx)
                if merged_val:
                    # Chỉ thêm vào path nếu chưa có (tránh lặp)
                    if not path or path[-1] != merged_val:
                        path.append(merged_val)
                continue
            
            val_str = str(val).strip()
            
            # Bỏ qua các giá trị rỗng hoặc ký tự đặc biệt
            if not val_str or val_str in ['nan', 'NaN', 'None']:
                continue
            
            # Thêm vào path nếu chưa có
            if not path or path[-1] != val_str:
                path.append(val_str)
        
        # Nếu path rỗng, đặt tên mặc định
        if not path:
            path = [f"Column_{col_idx}"]
        
        return path
    
    def _find_merged_value(self, row: List, col_idx: int) -> Optional[str]:
        """
        Tìm giá trị của merged cell bằng cách tìm ngược về bên trái.
        """
        
        for i in range(col_idx - 1, -1, -1):
            if pd.notna(row[i]):
                val = str(row[i]).strip()
                if val and val not in ['nan', 'NaN', 'None']:
                    return val
        
        return None
    
    def _parse_data_rows(self) -> List[Dict[str, Any]]:
        """Parse các hàng dữ liệu thành list of nested dictionaries."""
        
        data_rows = []
        
        for idx in range(self.data_start_row, len(self.df)):
            row = self.df.iloc[idx]
            
            # Kiểm tra hàng rỗng (tất cả đều NaN)
            if row.isna().all():
                continue
            
            row_data = self._parse_single_row(row)
            data_rows.append(row_data)
        
        return data_rows
    
    def _parse_single_row(self, row: pd.Series) -> Dict[str, Any]:
        """
        Parse một hàng dữ liệu thành nested dictionary dựa trên column_structure.
        """
        
        result = {}
        
        for col_info in self.column_structure:
            col_idx = col_info["col_index"]
            path = col_info["path"]
            value = self._safe_value(row[col_idx])
            
            # Xây dựng nested structure
            self._set_nested_value(result, path, value)
        
        return result
    
    def _set_nested_value(self, data: Dict, path: List[str], value: Any):
        """
        Đặt giá trị vào nested dictionary theo path.
        
        Ví dụ: path = ["Group1", "SubGroup", "Data"] 
               -> data["Group1"]["SubGroup"]["Data"] = value
        """
        
        if not path:
            return
        
        # Nếu path chỉ có 1 phần tử, gán trực tiếp
        if len(path) == 1:
            data[path[0]] = value
            return
        
        # Nếu path có nhiều phần tử, tạo nested structure
        current = data
        
        for i, key in enumerate(path[:-1]):
            if key not in current:
                current[key] = {}
            elif not isinstance(current[key], dict):
                # Xung đột: key đã tồn tại nhưng không phải dict
                # Chuyển thành dict và giữ giá trị cũ
                old_value = current[key]
                current[key] = {"_value": old_value}
            
            current = current[key]
        
        # Đặt giá trị cuối cùng
        final_key = path[-1]
        current[final_key] = value
    
    def _safe_value(self, val: Any) -> Any:
        """Chuyển đổi giá trị an toàn, xử lý NaN và kiểu dữ liệu."""
        
        if pd.isna(val):
            return None
        
        # Chuyển numpy types sang Python native types
        if hasattr(val, 'item'):
            val = val.item()
        
        # Xử lý số
        if isinstance(val, (int, float)):
            if isinstance(val, float):
                if val.is_integer():
                    return int(val)
            return val
        
        # Xử lý chuỗi
        val_str = str(val).strip()
        return val_str if val_str else None


def excel_to_nested_json(df: pd.DataFrame, 
                         output_file: Optional[str] = None,
                         indent: int = 2) -> Dict[str, Any]:
    """
    Chuyển đổi DataFrame với header nhiều cấp sang nested JSON.
    
    Function này hoàn toàn ĐỘNG - tự động phát hiện cấu trúc header.
    
    Parameters:
    -----------
    df : pd.DataFrame
        DataFrame đọc từ Excel với header=None
    output_file : str, optional
        Đường dẫn file JSON output. Nếu None, không ghi file.
    indent : int
        Số space cho indentation trong JSON
        
    Returns:
    --------
    dict : Nested JSON structure
    
    Example:
    --------
    >>> import pandas as pd
    >>> df = pd.read_excel('data.xlsx', header=None)
    >>> result = excel_to_nested_json(df, 'output.json')
    >>> print(json.dumps(result, indent=2, ensure_ascii=False))
    """
    
    parser = DynamicExcelParser(df)
    result = parser.parse()
    
    # Ghi file nếu được chỉ định
    if output_file:
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(result, f, indent=indent, ensure_ascii=False)
        print(f"✅ Đã lưu JSON vào: {output_file}")
        print(f"📊 Số hàng dữ liệu: {len(result['data'])}")
        print(f"📋 Số cột: {result['metadata']['total_columns']}")
    
    return result


def visualize_structure(result: Dict[str, Any]) -> None:
    """
    In ra cấu trúc cột để kiểm tra.
    """
    print("\n" + "="*80)
    print("CẤU TRÚC CỘT ĐƯỢC PHÁT HIỆN")
    print("="*80)
    
    for col in result['metadata']['column_structure']:
        print(f"Cột {col['col_index']:2d}: {col['full_path']}")
    
    print("\n" + "="*80)
    print(f"Tổng số cột: {result['metadata']['total_columns']}")
    print(f"Số hàng header: {result['metadata']['header_rows']}")
    print(f"Số hàng dữ liệu: {len(result['data'])}")
    print("="*80 + "\n")



# --- [GIAI ĐOẠN 2: PARSE LOGIC - HÀM MỚI] ---

def detect_header_split_point(
    raw_table_df: pd.DataFrame, 
    worksheet: Worksheet,
    boundary: Dict[str, int],
    border_threshold: float = 0.95  # Giá trị từ 0.0 đến 1.0
) -> int:
    """
    Phát hiện header bằng cách tìm ĐƯỜNG KẺ NGANG CUỐI CÙNG kéo dài suốt bảng.
    
    Logic:
    1. Bỏ qua hàng đầu tiên (viền trên của table)
    2. Quét TẤT CẢ các hàng và tìm hàng CUỐI CÙNG có kẻ ngang >= threshold
    3. Hàng đó là ranh giới header/data
    
    Args:
        border_threshold: Tỷ lệ từ 0.0 đến 1.0 (VD: 0.95 = 95%, 1.0 = 100%)
    
    Returns:
        Index (0-based) của hàng DATA đầu tiên. Trả về -1 nếu không tìm thấy.
    """
    
    # Validate threshold
    if not 0 <= border_threshold <= 1:
        print(f"⚠ CẢNH BÁO: border_threshold phải từ 0.0 đến 1.0, nhận được: {border_threshold}")
        border_threshold = max(0.0, min(1.0, border_threshold))
    
    total_columns = raw_table_df.shape[1]
    if total_columns == 0:
        return -1

    total_rows = raw_table_df.shape[0]
    if total_rows <= 1:
        return -1

    print(f"[detect_header_split_point] Quét {total_rows} hàng, {total_columns} cột")
    print(f"  Threshold: {border_threshold} ({border_threshold*100:.1f}%)")
    print(f"  Số cells tối thiểu: {int(border_threshold * total_columns)}/{total_columns}\n")

    last_border_row = -1
    
    # Quét từ hàng 1 (bỏ hàng 0 - viền trên)
    for r_idx in range(1, total_rows):
        real_row_num = boundary['min_row'] + r_idx
        horizontal_count = 0
        
        # Đếm cells có border TOP
        for c_idx in range(boundary['min_col'], boundary['max_col'] + 1):
            cell = worksheet.cell(row=real_row_num, column=c_idx)
            
            if cell.border.top and cell.border.top.style and cell.border.top.style != 'none':
                horizontal_count += 1
        
        border_rate = horizontal_count / total_columns
        
        # In kết quả
        status = ""
        if border_rate >= border_threshold:
            last_border_row = r_idx
            status = " ✓ ỨNG VIÊN"
        
        print(f"  Hàng {r_idx:2d} (Excel {real_row_num:2d}): "
              f"{horizontal_count:2d}/{total_columns:2d} = "
              f"{border_rate:5.1%}{status}")
    
        # Kết luận
        if last_border_row != -1:
            print(f"\n✓ Ranh giới tại hàng {last_border_row}")
            print(f"  Header: 0-{last_border_row-1}, Data: {last_border_row}+")
            return last_border_row
    
    print(f"\n✗ Không tìm thấy hàng nào >= {border_threshold*100:.0f}%")
    return -1


def detect_attribute_boundary(header_df: pd.DataFrame) -> Tuple[List[int], List[int]]:
    """
    (Hàm MỚI - Bước 2.5)
    Phân tích `header_df` (Cái Khuôn) để tìm "Ranh giới Thuộc tính".
    
    Quy tắc (Heuristic):
    - "Cột Thuộc tính" (Ngày, ID) chỉ có giá trị ở hàng đầu tiên (index 0).
    - "Cột Dữ liệu" (Group 1) có giá trị ở cả hàng 0 VÀ các hàng dưới.
    - Ranh giới là cột "Dữ liệu" đầu tiên được tìm thấy.
    
    Returns:
        Một tuple chứa 2 list: (attribute_cols_idx, data_cols_idx)
    """
    print(f"\n[detect_attribute_boundary] Phân tích {header_df.shape[1]} cột header...")
    
    attribute_cols_idx = []
    data_cols_idx = []
    
    total_header_rows = header_df.shape[0]
    total_cols = header_df.shape[1]

    # Trường hợp Bảng Đơn giản (header_df chỉ có 1 hàng)
    if total_header_rows == 1:
        print("  -> Phát hiện Bảng Đơn giản (1 hàng header).")
        # Giả định: Cột đầu tiên là Thuộc tính, còn lại là Dữ liệu
        attribute_cols_idx = [0]
        data_cols_idx = list(range(1, total_cols))
        
        print(f"  -> Cột Thuộc tính: {attribute_cols_idx}")
        print(f"  -> Cột Dữ liệu: {data_cols_idx}")
        return attribute_cols_idx, data_cols_idx

    # Trường hợp Bảng Phức tạp (header_df có > 1 hàng)
    print("  -> Phát hiện Bảng Phức tạp (>1 hàng header).")
    
    for c_idx in header_df.columns:
        # Lấy "thân" của cột (tất cả các hàng TRỪ hàng đầu tiên)
        column_body = header_df.iloc[1: , c_idx]
        
        # Kiểm tra xem "thân" có dữ liệu (không phải toàn NaN) không
        body_has_data = not column_body.isna().all()
        
        if body_has_data:
            # Đây là ranh giới! Cột này là "Cột Dữ liệu" đầu tiên.
            print(f"  -> Ranh giới tại Cột {c_idx} (vì có '{column_body.loc[column_body.notna().idxmax()]}')")
            
            # Tất cả các cột từ đây về sau ĐỀU LÀ Cột Dữ liệu
            data_cols_idx = list(range(c_idx, total_cols))
            
            # Thoát vòng lặp
            break
        else:
            # Nếu "thân" toàn NaN, đây là "Cột Thuộc tính"
            print(f"  -> Cột {c_idx} ('{header_df.iloc[0, c_idx]}') là Cột Thuộc tính.")
            attribute_cols_idx.append(c_idx)

    print(f"\n  -> [CHỐT] Cột Thuộc tính: {attribute_cols_idx}")
    print(f"  -> [CHỐT] Cột Dữ liệu: {data_cols_idx}")
    return attribute_cols_idx, data_cols_idx


# --- [GIAI ĐOẠN 3: Trích xuất JSON] ---


def _set_nested_value(target_dict: Dict, path: List[str], value: Any):
    """
    (Hàm trợ giúp - Bánh xe)
    Đi theo `path` và gán `value` ở cấp cuối cùng.
    Ví dụ: _set_nested_value(d, ['Group 1', 'Sub 1'], 5)
    -> d['Group 1']['Sub 1'] = 5
    """
    for key in path[:-1]:
        # Nếu key chưa có, tạo 1 dict con
        target_dict = target_dict.setdefault(key, {})
    # Gán giá trị ở cấp cuối cùng
    target_dict[path[-1]] = value


def _build_header_map(header_df: pd.DataFrame, data_cols: List[int]) -> Dict[int, List[str]]:
    """
    (Hàm MỚI - Bước 2.3)
    Phân tích `header_df` và tạo "Bản đồ Header" cho các cột dữ liệu.
    
    Logic:
    1. Lấp đầy (ffill) các ô gộp (cả ngang và dọc).
    2. Đọc "dọc" từng cột để xây dựng "con đường" (path).
    
    Returns:
        Một dict (bản đồ): { column_index -> [path, to, header] }
        Ví dụ: { 5: ['(Group 1)', 'Sub-Group 1.1', 'F-Data'] }
    """
    print(f"\n[build_header_map] Đang xây dựng bản đồ cho {len(data_cols)} cột dữ liệu...")
    
    # 1. Lấp đầy (ffill) để xử lý ô gộp
    # Fill ngang (axis=1) để vá các lỗ hổng ô gộp
    header_df_filled = header_df.ffill(axis=1)
    # Fill dọc (axis=0) để lấp đầy các cấp (ví dụ: Sub-Group 1.1)
    header_df_filled = header_df_filled.ffill(axis=0)
    
    header_map = {}
    
    # Chỉ lặp qua các CỘT DỮ LIỆU
    for c_idx in data_cols:
        path = []
        last_val = None # Dùng để tránh lặp lại (ví dụ: Group 1, Group 1, Group 1...)
        
        # Lặp qua từng hàng (row_index) trong header_df
        for r_idx in header_df_filled.index:
            value = header_df_filled.loc[r_idx, c_idx]
            
            # Chỉ thêm nếu nó không NaN VÀ không bị lặp lại
            if pd.notna(value) and value != last_val:
                path.append(value)
                last_val = value
        
        header_map[c_idx] = path
    
    # print(f"  -> Bản đồ Header (mẫu): Cột 5 -> {header_map.get(5)}")
    return header_map

def parse_table_to_long_json(
    header_df: pd.DataFrame, 
    data_df: pd.DataFrame, 
    attribute_cols: List[int], 
    data_cols: List[int]
) -> List[Dict[str, Any]]:
    """
    (Hàm MỚI - Bước 2.4)
    Lắp ráp JSON theo định dạng "Dài" (Long Format)
    (Một object JSON cho mỗi Ô dữ liệu).
    """
    
    final_json_list = []
    
    # --- 1. Chuẩn bị 2 "Bản đồ" ---
    
    # Bản đồ 1: "Bản đồ Header" (Tra cứu Path theo Cột)
    header_map = _build_header_map(header_df, data_cols)
    
    # Bản đồ 2: "Tên Thuộc tính" (Lấy tên "Ngày", "ID" từ hàng đầu)
    attribute_key_names = [header_df.iloc[0, c_idx] for c_idx in attribute_cols]
    
    print(f"[parse_table_to_long_json] Đang lắp ráp các ô...")

    # --- 2. Vòng lặp Kép (Lắp ráp Ô) ---
    
    # Lặp qua các HÀNG DỮ LIỆU (ví dụ: index 5, 6)
    for r_idx in data_df.index:
        
        # a. Lấy "Bản ghi Thuộc tính" (Attribute Record) cho hàng này
        # (Lấy 1 lần cho mỗi hàng)
        base_record = {}
        for i, c_idx in enumerate(attribute_cols):
            key = attribute_key_names[i]
            value = data_df.loc[r_idx, c_idx]
            base_record[key] = value
        
        # b. Lặp qua các CỘT DỮ LIỆU (ví dụ: 2, 3, ..., 25)
        for c_idx in data_cols:
            
            # i. Lấy Giá trị (Value)
            value = data_df.loc[r_idx, c_idx]
            
            # Bỏ qua nếu ô đó trống (không tạo JSON cho ô NaN)
            if pd.isna(value):
                continue
                
            # ii. Lấy "Con đường" (Path)
            path = header_map[c_idx]
            
            # iii. Lắp ráp
            
            # Tạo bản sao của "Bản ghi Thuộc tính"
            record = base_record.copy() 
            
            # Tạo object lồng nhau (Keys)
            nested_data_obj = {}
            _set_nested_value(nested_data_obj, path, value)
            
            # Gộp 2 phần lại
            record.update(nested_data_obj)
            
            # Thêm vào kết quả cuối cùng
            final_json_list.append(record)
            
    return final_json_list


if __name__ == "__main__":

# ======================================================================
# GIAI ĐOẠN TEST 1: PHÁT HIỆN BẢNG
# ======================================================================

    # # Thay đổi đường dẫn này cho đúng với file của bạn
    # # FILE_PATH = "path/to/your/image_b9d51d.xlsx"
    # FILE_PATH = "Book1.xlsx" # Giả sử file tên là report.xlsx
    # SHEET_NAME = "Sheet1"     # Thay tên sheet nếu cần

    # print(f"--- Bắt đầu phát hiện bảng trong file: {FILE_PATH} ---")
    
    # # Bạn có thể điều chỉnh 'min_width' và 'min_height'
    # table_coordinates = detect_tables(
    #     FILE_PATH, 
    #     SHEET_NAME, 
    #     min_width=2, 
    #     min_height=2
    # )
    
    # print("\n--- KẾT QUẢ CUỐI CÙNG ---")
    # if table_coordinates:
    #     for i, coords in enumerate(table_coordinates):
    #         print(f"Bảng {i+1} tìm thấy tại (1-indexed):")
    #         print(f"  - Hàng: từ {coords['min_row']} đến {coords['max_row']}")
    #         print(f"  - Cột:  từ {coords['min_col']} đến {coords['max_col']}")
    # else:
    #     print("Không tìm thấy bảng nào hợp lệ.")




# ======================================================================
# GIAI ĐOẠN TEST 2: PHÁT HIỆN BẢNG VÀ TRÍCH XUẤT DỮ LIỆU 
# ======================================================================

    # FILE_PATH = "basic_test.xlsx" # Sửa lại tên file của bạn
    # SHEET_NAME = "Sheet1"    # Sửa lại tên sheet của bạn

    # print(f"--- Bắt đầu phát hiện bảng trong file: {FILE_PATH} ---")
    
    # table_coordinates = detect_tables(
    #     FILE_PATH, 
    #     SHEET_NAME, 
    #     min_width=2,  # Giữ nguyên min_width=2, min_height=2 như bạn test
    #     min_height=2
    # )
    
    # print("\n--- KẾT QUẢ CUỐI CÙNG (PHÁT HIỆN) ---")
    # if table_coordinates:
    #     for i, coords in enumerate(table_coordinates):
    #         print(f"Bảng {i+1} tìm thấy tại (1-indexed):")
    #         print(f"  - Hàng: từ {coords['min_row']} đến {coords['max_row']}")
    #         print(f"  - Cột:  từ {coords['min_col']} đến {coords['max_col']}")
            
    #         # --- PHẦN DEBUG MỚI ---
    #         print(f"\n[DEBUG] Đang trích xuất dữ liệu thô Bảng {i+1}...")
    #         raw_data = debug_extract_data(FILE_PATH, SHEET_NAME, coords)
    #         print(type(raw_data))
    #         print(raw_data)
    #         if not raw_data.empty:
    #             print(f"--- Dữ liệu thô Bảng {i+1} (đầu & cuối): ---")
    #             # Hiển thị 5 hàng đầu và 5 hàng cuối của bảng
    #         #     with pd.option_context('display.max_rows', 10, 'display.max_columns', None):
    #         #         print(raw_data.head())
    #         # print("-" * 30)
    #         # --- KẾT THÚC PHẦN DEBUG --- 
            
    # else:
    #     print("Không tìm thấy bảng nào hợp lệ.")



# ======================================================================
# GIAI ĐOẠN TEST 3: PHÁT HIỆN HEADER VÀ TÁCH DỮ LIỆU 
# ======================================================================


    # FILE_PATH = "basic_test.xlsx" # File test của bạn
    # SHEET_NAME = "basic3"         # Sheet của bạn

    # # --- PHẢI LOAD `worksheet` TRƯỚC ---
    # try:
    #     wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    #     if SHEET_NAME not in wb.sheetnames:
    #         raise ValueError(f"Không tìm thấy sheet '{SHEET_NAME}'")
    #     worksheet = wb[SHEET_NAME]
    # except Exception as e:
    #     print(f"Lỗi khi tải workbook: {e}")
    #     exit()

    # # --- CHẠY GIAI ĐOẠN 1 (ĐỂ LẤY ĐẦU VÀO) ---
    # print(f"--- [GIAI ĐOẠN 1] Đang chạy detect_tables... ---")
    # table_coordinates = detect_tables(
    #     FILE_PATH, 
    #     SHEET_NAME, 
    #     min_width=2,
    #     min_height=2
    # )
    # print(f"--- [GIAI ĐOẠN 1] Hoàn thành: Tìm thấy {len(table_coordinates)} bảng ---")

    # # Lặp qua các bảng tìm được
    # for i, coords in enumerate(table_coordinates):
    #     print(f"\n--- Xử lý Bảng {i+1} (Hàng {coords['min_row']}->{coords['max_row']}) ---")
        
    #     raw_table_df = debug_extract_data(FILE_PATH, SHEET_NAME, coords)
        
    #     if raw_table_df.empty:
    #         continue
        
    #     # --- CHẠY HÀM MỚI (CHỈ DÙNG BORDER) ---
    #     split_point_index = detect_header_split_point(
    #         raw_table_df, 
    #         worksheet,   # Truyền worksheet
    #         coords,      # Truyền tọa độ
    #         border_threshold=0.95 # Chỉ truyền ngưỡng border
    #     )
    #     # ---
        
    #     if split_point_index != -1:
    #         # Kiểm tra trường hợp ranh giới vượt quá số hàng (hiếm gặp)
    #         if split_point_index >= len(raw_table_df.index):
    #              print(f"\n--- Kết quả Bảng {i+1}: Ranh giới ({split_point_index}) vượt quá số hàng. Bảng có thể chỉ có Header.")
    #              continue

    #         print(f"\n--- Kết quả Bảng {i+1} ---")
    #         print(f"  -> Ranh giới (Split Point) tìm thấy tại index hàng: {split_point_index}")
            
    #         header_df = raw_table_df.iloc[0 : split_point_index]
    #         data_df = raw_table_df.iloc[split_point_index : ]
            
    #         print("\n  -> [KHỐI HEADER] (Keys):")
    #         print(header_df)
    #         print("\n  -> [KHỐI DỮ LIỆU] (Values):")
    #         print(data_df)
    #         print("-" * 30)
            
    #     else:
    #         print(f"\n--- Kết quả Bảng {i+1}: Không thể xác định ranh giới Header/Data ---")

    # wb.close() # Đóng workbook sau khi xong


# ======================================================================
# GIAI ĐOẠN TEST 4: TÁCH THUỘC TÍNH TRONG HEADER VÀ IN KẾT QUẢ 
# ======================================================================

    
    # FILE_PATH = "basic_test.xlsx" # File test của bạn
    # SHEET_NAME = "basic3"         # Sheet của bạn

    # # --- PHẢI LOAD `worksheet` TRƯỚC ---
    # try:
    #     wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    #     if SHEET_NAME not in wb.sheetnames:
    #         raise ValueError(f"Không tìm thấy sheet '{SHEET_NAME}'")
    #     worksheet = wb[SHEET_NAME]
    # except Exception as e:
    #     print(f"Lỗi khi tải workbook: {e}")
    #     exit()

    # # --- CHẠY GIAI ĐOẠN 1 (ĐỂ LẤY ĐẦU VÀO) ---
    # print(f"--- [GIAI ĐOẠN 1] Đang chạy detect_tables... ---")
    # table_coordinates = detect_tables(
    #     FILE_PATH, 
    #     SHEET_NAME, 
    #     min_width=2,
    #     min_height=2
    # )
    # print(f"--- [GIAI ĐOẠN 1] Hoàn thành: Tìm thấy {len(table_coordinates)} bảng ---")

    # # Lặp qua các bảng tìm được
    # for i, coords in enumerate(table_coordinates):
    #     print(f"\n--- Xử lý Bảng {i+1} (Hàng {coords['min_row']}->{coords['max_row']}) ---")
        
    #     raw_table_df = debug_extract_data(FILE_PATH, SHEET_NAME, coords)
        
    #     if raw_table_df.empty:
    #         continue
        
    #     # --- BƯỚC 2.1: TÌM RANH GIỚI HEADER/DATA (Hàm của bạn) ---
    #     split_point_index = detect_header_split_point(
    #         raw_table_df, 
    #         worksheet,   # Truyền worksheet
    #         coords,      # Truyền tọa độ
    #         border_threshold=0.95 # Chỉ truyền ngưỡng border
    #     )
        
    #     if split_point_index != -1:
    #         if split_point_index >= len(raw_table_df.index):
    #              print(f"\n--- Kết quả Bảng {i+1}: Ranh giới ({split_point_index}) vượt quá số hàng.")
    #              continue

    #         print(f"\n--- Kết quả Bảng {i+1}: Tách Khối ---")
    #         print(f"  -> Ranh giới (Split Point) tìm thấy tại index hàng: {split_point_index}")
            
    #         header_df = raw_table_df.iloc[0 : split_point_index]
    #         data_df = raw_table_df.iloc[split_point_index : ]
            
    #         print("\n  -> [KHỐI HEADER] (Keys):")
    #         print(header_df.head()) # In 5 dòng đầu
            
    #         # --- BƯỚC 2.2: TÌM RANH GIỚI THUỘC TÍNH (Hàm MỚI) ---
    #         attribute_cols, data_cols = detect_attribute_boundary(header_df)
            
    #         print("-" * 30)
            
    #     else:
    #         print(f"\n--- Kết quả Bảng {i+1}: Không thể xác định ranh giới Header/Data ---")

    # wb.close() # Đóng workbook sau khi xong

# ======================================================================
# GIAI ĐOẠN TEST 5: CHẠY TOÀN BỘ VÀ IN KẾT QUẢ JSON
# ======================================================================

    
    FILE_PATH = "Book1.xlsx" # File test của bạn
    # SHEET_NAME = "basic3"         # Sheet của bạn
    SHEET_NAME = "Sheet1"         # Sheet của bạn

    # --- PHẢI LOAD `worksheet` TRƯỚC ---
    try:
        wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
        if SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"Không tìm thấy sheet '{SHEET_NAME}'")
        worksheet = wb[SHEET_NAME]
    except Exception as e:
        print(f"Lỗi khi tải workbook: {e}")
        exit()

    # --- CHẠY GIAI ĐOẠN 1 (ĐỂ LẤY ĐẦU VÀO) ---
    print(f"--- [GIAI ĐOẠN 1] Đang chạy detect_tables... ---")
    table_coordinates = detect_tables(
        FILE_PATH, 
        SHEET_NAME, 
        min_width=2,
        min_height=2
    )
    print(f"--- [GIAI ĐOẠN 1] Hoàn thành: Tìm thấy {len(table_coordinates)} bảng ---")

    all_parsed_data = [] # List cuối cùng chứa TẤT CẢ JSON

    # Lặp qua các bảng tìm được
    for i, coords in enumerate(table_coordinates):
        print(f"\n--- Xử lý Bảng {i+1} (Hàng {coords['min_row']}->{coords['max_row']}) ---")
        
        raw_table_df = debug_extract_data(FILE_PATH, SHEET_NAME, coords)
        
        if raw_table_df.empty:
            continue
        
        # --- BƯỚC 2.1: TÌM RANH GIỚI HEADER/DATA (Hàm của bạn) ---
        split_point_index = detect_header_split_point(
            raw_table_df, 
            worksheet,
            coords,
            border_threshold=0.95
        )
        
        if split_point_index != -1:
            if split_point_index >= len(raw_table_df.index):
                 print(f"\n--- Kết quả Bảng {i+1}: Ranh giới ({split_point_index}) vượt quá số hàng.")
                 continue

            header_df = raw_table_df.iloc[0 : split_point_index]
            data_df = raw_table_df.iloc[split_point_index : ]
            
            # --- BƯỚC 2.2: TÌM RANH GIỚI THUỘC TÍNH (Hàm MỚI) ---
            attribute_cols, data_cols = detect_attribute_boundary(header_df)
            
            # --- BƯỚC 2.3 & 2.4: LẮP RÁP JSON ---
            try:
                # Chạy hàm parse JSON (Định dạng "Dài")
                json_output = parse_table_to_long_json(
                    header_df, 
                    data_df, 
                    attribute_cols, 
                    data_cols
                )
                
                all_parsed_data.extend(json_output)
                print(f"\n--- [GIAI ĐOẠN 2] Parse Bảng {i+1} thành công. Tạo ra {len(json_output)} bản ghi JSON.")

            except Exception as e:
                print(f"LỖI khi parse Bảng {i+1}: {e}")
                import traceback
                traceback.print_exc()

            print("-" * 30)
            
        else:
            print(f"\n--- Kết quả Bảng {i+1}: Không thể xác định ranh giới Header/Data ---")

    wb.close() # Đóng workbook sau khi xong
    
    print("\n--- [HOÀN THÀNH] Đã xử lý tất cả các bảng. ---")
    
    # In toàn bộ kết quả cuối cùng
    print("\n--- TỔNG KẾT JSON ---")
    json_response = json.dumps(all_parsed_data, indent=2, ensure_ascii=False)

    ## Save to file
    OUTPUT_FILE = "final_output_test.json"
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write(json_response)
    print(f"✅ Đã lưu kết quả JSON vào: {OUTPUT_FILE}")